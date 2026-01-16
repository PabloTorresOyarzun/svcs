from litestar import Litestar, get, post, delete, Router, Request
from litestar.exceptions import HTTPException
from litestar.datastructures import UploadFile
from litestar.enums import RequestEncodingType
from litestar.params import Body
from litestar.openapi import OpenAPIConfig
from litestar.openapi.spec import Components, SecurityScheme
from litestar.status_codes import HTTP_200_OK, HTTP_400_BAD_REQUEST, HTTP_401_UNAUTHORIZED, HTTP_403_FORBIDDEN, HTTP_404_NOT_FOUND, HTTP_408_REQUEST_TIMEOUT, HTTP_413_REQUEST_ENTITY_TOO_LARGE, HTTP_500_INTERNAL_SERVER_ERROR
from pydantic import BaseModel
import httpx
import os
import base64
import io
import fitz
import sys
import asyncio
import logging
import time
from concurrent.futures import ThreadPoolExecutor
from contextlib import contextmanager
from typing import Optional, Dict, List, Any, Annotated
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type
import pandas as pd
from PIL import Image
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader

from quality import paso_1_analizar_documento, paso_2_corregir_rotacion
from clasificacion import clasificar_documento_completo, segmentar_pdf
from config import get_settings, calcular_timeout_excel, calcular_timeout_calidad, get_valid_api_tokens
from token_manager import token_manager
from database import db_manager, cache_repo, calcular_hash_documentos, calcular_hash_archivo


@contextmanager
def suprimir_prints():
    """Suprime temporalmente la salida a stdout."""
    original_stdout = sys.stdout
    sys.stdout = io.StringIO()
    try:
        yield
    finally:
        sys.stdout = original_stdout


try:
    pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
    FUENTE_PRINCIPAL = 'STSong-Light'
except Exception:
    FUENTE_PRINCIPAL = 'Helvetica'


logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

settings = get_settings()


def get_bearer_token(request: Request) -> str:
    """Extrae el token Bearer del header Authorization."""
    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        raise HTTPException(status_code=HTTP_401_UNAUTHORIZED, detail="Token de autenticación requerido")
    return auth_header[7:]


def verify_api_token(request: Request) -> str:
    """Verifica que el token API proporcionado sea válido."""
    token = get_bearer_token(request)
    
    if settings.ADMIN_TOKEN and token == settings.ADMIN_TOKEN:
        return token

    if token_manager.is_valid_token(token):
        return token

    valid_tokens = get_valid_api_tokens()
    if valid_tokens and token in valid_tokens:
        return token

    raise HTTPException(status_code=HTTP_401_UNAUTHORIZED, detail="Token de autenticación inválido")


def verify_admin_token(request: Request) -> str:
    """Verifica que el token de administrador sea válido."""
    token = get_bearer_token(request)
    
    if not settings.ADMIN_TOKEN:
        raise HTTPException(status_code=HTTP_500_INTERNAL_SERVER_ERROR, detail="ADMIN_TOKEN no configurado en el servidor")

    if token != settings.ADMIN_TOKEN:
        raise HTTPException(status_code=HTTP_403_FORBIDDEN, detail="Se requiere token de administrador")

    return token


BASE_URL = "https://backend.juanleon.cl"
ENDPOINT_DESPACHO = "/api/admin/despachos/{codigo}"
ENDPOINT_DOCUMENTOS = "/api/admin/documentos64/despacho/{codigo_visible}"

# Azure Document Intelligence - Cloud
API_VERSION = "2024-11-30"

# Modelo custom en Azure Cloud
CUSTOM_MODEL_ID = "master-01-alpha"

# Mapeo de tipos de documento al modelo (todos usan el mismo modelo master)
DOCUMENT_TYPE_TO_MODEL = {
    "FACTURA_COMERCIAL": CUSTOM_MODEL_ID,
    "DOCUMENTO_TRANSPORTE": CUSTOM_MODEL_ID,
    "CERTIFICADO_ORIGEN": CUSTOM_MODEL_ID,
    "LISTA_EMBALAJE": CUSTOM_MODEL_ID,
    "CERTIFICADO_SANITARIO": CUSTOM_MODEL_ID,
    "POLIZA_SEGURO": CUSTOM_MODEL_ID,
    "UNKNOWN_DOCUMENT": None
}

documentos_finales_sgd: Dict[str, List[Dict]] = {}
documentos_finales_individuales: Dict[str, List[Dict]] = {}

executor = ThreadPoolExecutor(max_workers=min(settings.EXECUTOR_MAX_WORKERS, (os.cpu_count() or 1) * 4))
logger.info(f"ThreadPoolExecutor inicializado con max_workers={executor._max_workers}")

MAX_CONCURRENT_PDFS = 10
pdf_semaphore = asyncio.Semaphore(MAX_CONCURRENT_PDFS)


# Pydantic Models
class DocumentoSimplificado(BaseModel):
    nombre: str
    estado: str
    fecha_recepcion: str


class Usuarios(BaseModel):
    pedidor: Optional[List[str]] = None
    jefe_operaciones: Optional[List[str]] = None


class ConsultaResponse(BaseModel):
    codigo_despacho: str
    id_interno: str
    cliente: str
    estado: str
    tipo: str
    total_documentos: int
    documentos: List[DocumentoSimplificado]
    usuarios: Usuarios


class Alerta(BaseModel):
    pagina: int
    tipo: str
    descripcion: str


class DocumentoFinal(BaseModel):
    archivo_origen: str
    nombre_salida: str
    tipo: str
    paginas: List[int]
    alertas: Optional[List[Alerta]] = None
    datos_extraidos: Optional[Dict[str, Any]] = None


class CacheInfo(BaseModel):
    desde_cache: bool
    hash_documentos: Optional[str] = None
    fecha_cache: Optional[str] = None
    hay_cambios: Optional[bool] = None


class ProcesamientoResponse(BaseModel):
    codigo_despacho: str
    cliente: str
    estado: str
    tipo: str
    total_documentos_segmentados: int
    documentos: List[DocumentoFinal]
    cache_info: Optional[CacheInfo] = None


class ProcesamientoIndividualResponse(BaseModel):
    archivo_origen: str
    total_documentos_segmentados: int
    documentos: List[DocumentoFinal]
    cache_info: Optional[CacheInfo] = None


class TokenInfo(BaseModel):
    id: str
    name: str
    masked_token: str
    created_at: str
    created_by: str
    last_used: Optional[str] = None
    is_active: bool = True


class TokenCreateRequest(BaseModel):
    name: str


class TokenCreateResponse(BaseModel):
    id: str
    token: str
    name: str
    created_at: str
    message: str


class TokenDeleteResponse(BaseModel):
    success: bool
    message: str


def validar_pdf(file_bytes: bytes) -> bool:
    """Valida que los bytes sean un PDF válido."""
    if not file_bytes.startswith(b'%PDF'):
        return False
    try:
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        doc.close()
        return True
    except Exception:
        return False


def validar_excel(file_bytes: bytes, nombre_archivo: str) -> bool:
    """Valida que los bytes sean un archivo Excel válido."""
    extension = os.path.splitext(nombre_archivo.lower())[1]
    
    if extension in ['.xlsx', '.xlsm', '.xltx', '.xltm']:
        if not file_bytes.startswith(b'PK\x03\x04'):
            return False
    elif extension in ['.xls', '.xlsb']:
        if not file_bytes.startswith(b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1'):
            return False
    
    try:
        engine = 'xlrd' if extension == '.xls' else 'openpyxl'
        pd.read_excel(io.BytesIO(file_bytes), engine=engine, nrows=0)
        return True
    except Exception:
        return False


def validar_tamano_archivo(file_bytes: bytes) -> None:
    """Valida que el archivo no exceda el tamaño máximo."""
    file_size_mb = len(file_bytes) / (1024 * 1024)
    if file_size_mb > settings.MAX_FILE_SIZE_MB:
        raise HTTPException(
            status_code=HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"Archivo excede el tamaño máximo de {settings.MAX_FILE_SIZE_MB}MB"
        )


@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=2, max=10),
    retry=retry_if_exception_type((httpx.TimeoutException, httpx.NetworkError))
)
async def consultar_despacho_detalle(codigo_interno: str, token: str) -> Optional[Dict]:
    if not token:
        return None
    
    url = f"{BASE_URL}{ENDPOINT_DESPACHO.format(codigo=str(codigo_interno))}"
    headers = {
        "Accept": "application/json",
        "Authorization": f"Bearer {token}"
    }
    
    timeout = httpx.Timeout(
        connect=settings.TIMEOUT_CONNECT,
        read=settings.TIMEOUT_READ,
        write=settings.TIMEOUT_WRITE,
        pool=5.0
    )
    
    try:
        async with httpx.AsyncClient(timeout=timeout) as client:
            response = await client.get(url, headers=headers)
            response.raise_for_status()
            
            if 'application/json' not in response.headers.get('Content-Type', ''):
                return None
            
            datos = response.json()
            return datos if isinstance(datos, dict) else None
    except (httpx.TimeoutException, httpx.NetworkError):
        raise
    except Exception:
        return None


@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=1, min=2, max=10),
    retry=retry_if_exception_type((httpx.TimeoutException, httpx.NetworkError))
)
async def consultar_documentacion(codigo: str, token: str) -> Optional[List]:
    if not token:
        return None
    
    url = f"{BASE_URL}{ENDPOINT_DOCUMENTOS.format(codigo_visible=str(codigo))}"
    headers = {
        "Accept": "application/json",
        "Authorization": f"Bearer {token}"
    }
    
    timeout = httpx.Timeout(
        connect=settings.TIMEOUT_CONNECT,
        read=settings.TIMEOUT_READ,
        write=settings.TIMEOUT_WRITE,
        pool=5.0
    )
    
    try:
        async with httpx.AsyncClient(timeout=timeout) as client:
            response = await client.get(url, headers=headers)
            response.raise_for_status()
            
            if 'application/json' not in response.headers.get('Content-Type', ''):
                return None
            
            datos_json = response.json()
            if not isinstance(datos_json, dict):
                return None
            
            documentos = datos_json.get("data", [])
            return documentos if isinstance(documentos, list) else None
    except (httpx.TimeoutException, httpx.NetworkError):
        raise
    except Exception:
        return None


def es_archivo_excel(nombre_archivo: str) -> bool:
    """Verifica si un archivo es Excel basándose en su extensión."""
    extensiones_excel = ['.xls', '.xlsx', '.xlsm', '.xlsb', '.xltx', '.xltm']
    extension = os.path.splitext(nombre_archivo.lower())[1]
    return extension in extensiones_excel


def es_archivo_imagen(nombre_archivo: str) -> bool:
    """Verifica si un archivo es una imagen basándose en su extensión."""
    extensiones_imagen = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif', '.webp']
    extension = os.path.splitext(nombre_archivo.lower())[1]
    return extension in extensiones_imagen


def validar_imagen(file_bytes: bytes, nombre_archivo: str) -> bool:
    """Valida que los bytes sean una imagen válida."""
    try:
        imagen = Image.open(io.BytesIO(file_bytes))
        imagen.verify()
        return True
    except Exception:
        return False


def formatear_celda(valor):
    """Formatea el valor de una celda para su representación en el PDF."""
    if pd.isna(valor):
        return ""

    if isinstance(valor, (float, int)):
        try:
            val_float = float(valor)
            if val_float.is_integer():
                return str(int(val_float))
            return str(val_float)
        except Exception:
            return str(valor)

    if isinstance(valor, pd.Timestamp):
        return valor.strftime('%Y-%m-%d')

    return str(valor)


def limpiar_dataframe(df):
    """Limpia el DataFrame eliminando columnas y filas vacías, detectando gaps."""
    df = df.fillna('')
    df = df.astype(str)
    df = df.replace('nan', '')

    densidades = []
    for col in df.columns:
        celdas_con_dato = (df[col].str.strip().str.len() > 0).sum()
        densidades.append(celdas_con_dato)

    umbral = 2

    hay_gap = False
    gap_actual = 0
    for d in densidades:
        if d < umbral:
            gap_actual += 1
            if gap_actual >= 3:
                hay_gap = True
                break
        else:
            gap_actual = 0

    if hay_gap:
        cols_con_datos_idx = [i for i, d in enumerate(densidades) if d >= umbral]

        if not cols_con_datos_idx:
            return pd.DataFrame()

        primer_col = cols_con_datos_idx[0]
        fin_bloque = cols_con_datos_idx[-1]

        for i in range(primer_col, len(densidades)):
            gap = 0
            for j in range(i, min(i + 5, len(densidades))):
                if densidades[j] < umbral:
                    gap += 1
                else:
                    break
            if gap >= 3:
                fin_bloque = i - 1
                break

        cols_seleccionadas = [i for i in range(primer_col, fin_bloque + 1) if densidades[i] >= umbral]
        if not cols_seleccionadas:
            return pd.DataFrame()
        df = df.iloc[:, cols_seleccionadas]
    else:
        cols_con_datos = []
        for col in df.columns:
            if df[col].str.strip().str.len().sum() > 0:
                cols_con_datos.append(col)
        df = df[cols_con_datos]

    df = df[df.apply(lambda row: row.str.strip().str.len().sum() > 0, axis=1)]

    return df.reset_index(drop=True)


def calcular_anchos_columnas_mejorado(df, disponible_width):
    """Calcula anchos de columnas óptimos basados en el contenido."""
    if df.empty:
        return []

    num_cols = len(df.columns)
    max_lens = []

    for col_idx in range(num_cols):
        col_data = df.iloc[:, col_idx].astype(str)
        longitudes = col_data.str.len()
        p90 = longitudes.quantile(0.9) if len(longitudes) > 0 else 0
        max_lens.append(max(p90, 5))

    total_chars = sum(max_lens)
    if total_chars == 0:
        total_chars = 1

    col_widths = [(l / total_chars) * disponible_width for l in max_lens]

    if num_cols <= 6:
        min_width = 0.8 * inch
        max_width = 5.0 * inch
    elif num_cols <= 10:
        min_width = 0.5 * inch
        max_width = 4.0 * inch
    else:
        min_width = 0.4 * inch
        max_width = 3.0 * inch

    col_widths = [max(min(w, max_width), min_width) for w in col_widths]

    current_total = sum(col_widths)
    if current_total > 0:
        factor = disponible_width / current_total
        col_widths = [w * factor for w in col_widths]

    return col_widths


def _convertir_imagen_a_pdf_sync(imagen_bytes: bytes, nombre_archivo: str) -> bytes:
    """Función síncrona interna para conversión de imagen a PDF."""
    try:
        imagen = Image.open(io.BytesIO(imagen_bytes))

        if imagen.mode == 'RGBA':
            fondo = Image.new('RGB', imagen.size, (255, 255, 255))
            fondo.paste(imagen, mask=imagen.split()[3])
            imagen = fondo
        elif imagen.mode != 'RGB':
            imagen = imagen.convert('RGB')

        ancho_img, alto_img = imagen.size
        ancho_pagina, alto_pagina = A4
        margen = 30

        ancho_disponible = ancho_pagina - (2 * margen)
        alto_disponible = alto_pagina - (2 * margen)

        escala = min(ancho_disponible / ancho_img, alto_disponible / alto_img)

        ancho_final = ancho_img * escala
        alto_final = alto_img * escala

        x = (ancho_pagina - ancho_final) / 2
        y = (alto_pagina - alto_final) / 2

        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)

        img_reader = ImageReader(imagen)
        c.drawImage(img_reader, x, y, width=ancho_final, height=alto_final, preserveAspectRatio=True)
        c.save()

        return buffer.getvalue()

    except Exception as e:
        raise HTTPException(
            status_code=HTTP_400_BAD_REQUEST,
            detail=f"Error al convertir imagen a PDF: {str(e)}"
        )


def _convertir_excel_a_pdf_sync(excel_bytes: bytes, nombre_archivo: str) -> bytes:
    """Función síncrona interna para conversión Excel a PDF con soporte avanzado."""
    try:
        extension = os.path.splitext(nombre_archivo.lower())[1]

        df_dict = None
        errores = []

        if extension in ['.xlsx', '.xlsm']:
            try:
                df_dict = pd.read_excel(io.BytesIO(excel_bytes), sheet_name=None, engine='openpyxl', header=None)
            except Exception as e1:
                errores.append(f"openpyxl: {str(e1)[:100]}")
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)
                    df_dict = {}
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        data = []
                        for row in ws.iter_rows(values_only=True):
                            data.append(list(row))
                        if data:
                            df_dict[sheet_name] = pd.DataFrame(data)
                except Exception as e2:
                    errores.append(f"openpyxl manual: {str(e2)[:100]}")
        elif extension == '.xls':
            try:
                df_dict = pd.read_excel(io.BytesIO(excel_bytes), sheet_name=None, engine='xlrd', header=None)
            except Exception as e:
                errores.append(f"xlrd: {str(e)[:100]}")
        else:
            try:
                df_dict = pd.read_excel(io.BytesIO(excel_bytes), sheet_name=None, engine='openpyxl', header=None)
            except Exception as e:
                errores.append(f"openpyxl: {str(e)[:100]}")

        if not df_dict:
            logger.warning(f"No se pudo leer Excel {nombre_archivo}: {'; '.join(errores)}")
            raise HTTPException(
                status_code=HTTP_400_BAD_REQUEST,
                detail=f"Error al leer Excel: {'; '.join(errores)}"
            )

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=0.3*inch,
            rightMargin=0.3*inch,
            topMargin=0.4*inch,
            bottomMargin=0.4*inch
        )

        elements = []
        styles = getSampleStyleSheet()

        estilo_titulo = ParagraphStyle(
            'TituloHoja',
            parent=styles['Heading1'],
            fontName=FUENTE_PRINCIPAL,
            fontSize=14,
            spaceAfter=12
        )

        estilo_celda = ParagraphStyle(
            'CeldaTabla',
            fontName=FUENTE_PRINCIPAL,
            fontSize=9,
            leading=11,
        )

        for sheet_name, df in df_dict.items():
            elements.append(Paragraph(f"Hoja: {sheet_name}", estilo_titulo))

            df_clean = limpiar_dataframe(df)

            if df_clean.empty:
                continue

            data = []
            for row in df_clean.values.tolist():
                fila_procesada = []
                for c in row:
                    texto = formatear_celda(c)
                    fila_procesada.append(Paragraph(texto, estilo_celda))
                data.append(fila_procesada)

            if not data:
                continue

            ancho_util = landscape(A4)[0] - 0.6*inch
            col_widths = calcular_anchos_columnas_mejorado(df_clean, ancho_util)

            table = Table(data, colWidths=col_widths, repeatRows=1)

            estilo_grid = [
                ('FONTNAME', (0, 0), (-1, -1), FUENTE_PRINCIPAL),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.93, 0.93, 0.93)),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ]

            table.setStyle(TableStyle(estilo_grid))

            elements.append(table)
            elements.append(PageBreak())

        doc.build(elements)
        return buffer.getvalue()

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=HTTP_400_BAD_REQUEST,
            detail=f"Error al convertir Excel a PDF: {str(e)}"
        )


async def convertir_excel_a_pdf(excel_bytes: bytes, nombre_archivo: str) -> bytes:
    """Convierte un archivo Excel a PDF de forma asíncrona con timeout dinámico."""
    timeout = calcular_timeout_excel(len(excel_bytes))
    file_size_mb = len(excel_bytes) / (1024 * 1024)

    logger.info(f"Iniciando conversión Excel a PDF para {nombre_archivo} ({file_size_mb:.2f}MB, timeout={timeout}s)")
    start_time = time.time()

    loop = asyncio.get_event_loop()

    try:
        result = await asyncio.wait_for(
            loop.run_in_executor(
                executor,
                _convertir_excel_a_pdf_sync,
                excel_bytes,
                nombre_archivo
            ),
            timeout=timeout
        )

        elapsed_time = time.time() - start_time
        logger.info(f"Conversión Excel completada para {nombre_archivo} en {elapsed_time:.2f}s")

        return result
    except asyncio.TimeoutError:
        logger.error(f"Conversión Excel excedió timeout para {nombre_archivo} ({timeout}s)")
        raise HTTPException(
            status_code=HTTP_408_REQUEST_TIMEOUT,
            detail=f"Conversión Excel excedió el tiempo límite de {timeout}s"
        )


async def convertir_imagen_a_pdf(imagen_bytes: bytes, nombre_archivo: str) -> bytes:
    """Convierte una imagen a PDF de forma asíncrona."""
    file_size_mb = len(imagen_bytes) / (1024 * 1024)
    timeout = 60

    logger.info(f"Iniciando conversión imagen a PDF para {nombre_archivo} ({file_size_mb:.2f}MB)")
    start_time = time.time()

    loop = asyncio.get_event_loop()

    try:
        result = await asyncio.wait_for(
            loop.run_in_executor(
                executor,
                _convertir_imagen_a_pdf_sync,
                imagen_bytes,
                nombre_archivo
            ),
            timeout=timeout
        )

        elapsed_time = time.time() - start_time
        logger.info(f"Conversión imagen completada para {nombre_archivo} en {elapsed_time:.2f}s")

        return result
    except asyncio.TimeoutError:
        logger.error(f"Conversión imagen excedió timeout para {nombre_archivo} ({timeout}s)")
        raise HTTPException(
            status_code=HTTP_408_REQUEST_TIMEOUT,
            detail=f"Conversión imagen excedió el tiempo límite de {timeout}s"
        )


def get_azure_headers() -> Dict[str, str]:
    """Retorna headers para autenticación con Azure DI Cloud."""
    return {
        "Ocp-Apim-Subscription-Key": settings.AZURE_KEY,
        "Content-Type": "application/pdf"
    }


def get_azure_base_url() -> str:
    """Retorna la URL base de Azure DI."""
    endpoint = settings.AZURE_ENDPOINT.rstrip('/')
    return f"{endpoint}/documentintelligence"


def limpiar_datos_azure(data: Any) -> Any:
    """
    Limpia recursivamente la respuesta de Azure DI para eliminar metadata
    (polygons, spans, confidence) y dejar solo los valores.
    """
    if isinstance(data, list):
        return [limpiar_datos_azure(item) for item in data]
    
    if isinstance(data, dict):
        if "valueString" in data: return data["valueString"]
        if "valueNumber" in data: return data["valueNumber"]
        if "valueDate" in data: return data["valueDate"]
        if "valueTime" in data: return data["valueTime"]
        if "valuePhoneNumber" in data: return data["valuePhoneNumber"]
        if "valueBoolean" in data: return data["valueBoolean"]
        if "valueSelectionMark" in data: return data["valueSelectionMark"]
        
        if "valueArray" in data:
            return limpiar_datos_azure(data["valueArray"])
        if "valueObject" in data:
            return {k: limpiar_datos_azure(v) for k, v in data["valueObject"].items()}
        
        cleaned = {}
        for k, v in data.items():
            if k in ["boundingRegions", "polygon", "spans", "confidence", "type"]:
                continue
            cleaned[k] = limpiar_datos_azure(v)
        
        if not cleaned and "content" in data:
            return data["content"]
            
        return cleaned

    return data


def eliminar_campos_vacios(data: Any) -> Any:
    """
    Elimina recursivamente claves con valores None, {}, [] o strings vacíos.
    """
    if isinstance(data, dict):
        cleaned = {
            k: eliminar_campos_vacios(v)
            for k, v in data.items()
        }
        return {
            k: v for k, v in cleaned.items() 
            if v not in [None, {}, [], ""]
        }
    
    elif isinstance(data, list):
        cleaned = [eliminar_campos_vacios(item) for item in data]
        return [item for item in cleaned if item not in [None, {}, [], ""]]
        
    return data


async def verificar_modelo_entrenado(model_id: str) -> bool:
    """Verifica si un modelo custom está entrenado en Azure DI Cloud."""
    base_url = get_azure_base_url()
    url = f"{base_url}/documentModels/{model_id}?api-version={API_VERSION}"

    headers = {
        "Ocp-Apim-Subscription-Key": settings.AZURE_KEY
    }

    timeout_config = httpx.Timeout(
        connect=settings.TIMEOUT_CONNECT,
        read=settings.TIMEOUT_READ,
        write=settings.TIMEOUT_WRITE,
        pool=5.0
    )

    try:
        async with httpx.AsyncClient(timeout=timeout_config) as client:
            response = await client.get(url, headers=headers)
            if response.status_code == 200:
                logger.info(f"Modelo {model_id} verificado en Azure Cloud")
                return True
            else:
                logger.warning(f"Modelo {model_id} no encontrado: {response.status_code}")
                return False
    except Exception as e:
        logger.error(f"Error verificando modelo {model_id}: {str(e)}")
        return False


async def extraer_datos_con_modelo(pdf_bytes: bytes, model_id: str) -> Optional[Dict[str, Any]]:
    """Extrae datos estructurados de un PDF usando un modelo custom de Azure DI Cloud."""
    base_url = get_azure_base_url()
    url = f"{base_url}/documentModels/{model_id}:analyze?api-version={API_VERSION}"

    headers = get_azure_headers()

    timeout_config = httpx.Timeout(
        connect=settings.TIMEOUT_CONNECT,
        read=300.0,
        write=settings.TIMEOUT_WRITE,
        pool=5.0
    )

    try:
        async with httpx.AsyncClient(timeout=timeout_config) as client:
            logger.info(f"Enviando documento a modelo {model_id} en Azure Cloud")
            response = await client.post(url, headers=headers, content=pdf_bytes)

            if response.status_code != 202:
                logger.warning(f"Error iniciando análisis: {response.status_code} - {response.text[:500]}")
                return None

            operation_location = response.headers.get('Operation-Location')
            if not operation_location:
                logger.warning("No se recibió Operation-Location")
                return None

            logger.info(f"Análisis iniciado, polling: {operation_location}")

            poll_headers = {
                "Ocp-Apim-Subscription-Key": settings.AZURE_KEY
            }

            max_intentos = 60
            for intento in range(max_intentos):
                status_response = await client.get(operation_location, headers=poll_headers)

                if status_response.status_code != 200:
                    logger.warning(f"Error en polling: {status_response.status_code}")
                    return None

                resultado = status_response.json()
                status = resultado.get('status')

                if status == 'succeeded':
                    analyze_result = resultado.get('analyzeResult', {})
                    documentos = analyze_result.get('documents', [])
                    
                    logger.info(f"Análisis exitoso. Documentos encontrados: {len(documentos)}")

                    if documentos:
                        fields = documentos[0].get('fields', {})
                        logger.info(f"Campos extraídos: {list(fields.keys())}")

                        datos_aplanados = {
                            k: limpiar_datos_azure(v) for k, v in fields.items()
                        }

                        datos_finales = eliminar_campos_vacios(datos_aplanados)

                        return datos_finales
                    return {}

                elif status == 'failed':
                    error = resultado.get('error', {})
                    logger.error(f"Análisis fallido: {error}")
                    return None

                await asyncio.sleep(2)

            logger.warning("Timeout en análisis")
            return None

    except Exception as e:
        logger.error(f"Error en extracción: {str(e)}")
        return None


def _procesar_calidad_sync(pdf_bytes: bytes) -> tuple:
    """Función síncrona interna para procesamiento de calidad."""
    pdf_doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    
    with suprimir_prints():
        resultados_paso_1 = paso_1_analizar_documento(pdf_doc)
        paginas_corregidas = paso_2_corregir_rotacion(pdf_doc, resultados_paso_1)
    
    if paginas_corregidas > 0:
        pdf_bytes = pdf_doc.tobytes()
    
    pdf_doc.close()
    return pdf_bytes, resultados_paso_1


async def clasificar_pdf_completo(pdf_bytes: bytes, nombre_archivo: str) -> Dict[str, Any]:
    """Flujo de clasificación de PDF (sin extracción de datos)."""
    timeout_calidad = 0
    try:
        pdf_doc_temp = fitz.open(stream=pdf_bytes, filetype="pdf")
        num_paginas = len(pdf_doc_temp)
        pdf_doc_temp.close()

        timeout_calidad = calcular_timeout_calidad(num_paginas)

        logger.info(f"Iniciando procesamiento de calidad para {nombre_archivo} ({num_paginas} páginas, timeout={timeout_calidad}s)")
        start_time = time.time()

        loop = asyncio.get_event_loop()
        pdf_bytes, resultados_paso_1 = await asyncio.wait_for(
            loop.run_in_executor(
                executor,
                _procesar_calidad_sync,
                pdf_bytes
            ),
            timeout=timeout_calidad
        )

        elapsed_time = time.time() - start_time
        logger.info(f"Procesamiento de calidad completado para {nombre_archivo} en {elapsed_time:.2f}s")

        clasificaciones = await clasificar_documento_completo(pdf_bytes)

        logger.info(f"Iniciando segmentación para {nombre_archivo}")
        start_time = time.time()

        documentos_segmentados = await loop.run_in_executor(
            executor,
            segmentar_pdf,
            pdf_bytes,
            clasificaciones
        )

        elapsed_time = time.time() - start_time
        logger.info(f"Segmentación completada para {nombre_archivo} en {elapsed_time:.2f}s - {len(documentos_segmentados)} documentos")

        alertas_por_documento = {}
        for resultado in resultados_paso_1:
            alertas = []

            if resultado['escaneada']:
                if 'INCLINADA' in resultado['orientacion']:
                    alertas.append({
                        "pagina": resultado['pagina'],
                        "tipo": "inclinado",
                        "descripcion": f"Página escaneada {resultado['orientacion']}"
                    })
                elif resultado['orientacion'] not in ['NORMAL', 'SIN TEXTO', 'SIN IMAGEN']:
                    alertas.append({
                        "pagina": resultado['pagina'],
                        "tipo": "escaneado",
                        "descripcion": f"Página escaneada: {resultado['orientacion']}"
                    })

            if resultado['rotacion_formal'] != 0:
                alertas.append({
                    "pagina": resultado['pagina'],
                    "tipo": "rotado",
                    "descripcion": f"Rotación de {resultado['rotacion_formal']}° corregida"
                })

            if not resultado['escaneada'] and resultado['orientacion'] == 'ROTADA':
                alertas.append({
                    "pagina": resultado['pagina'],
                    "tipo": "rotado",
                    "descripcion": "Texto vertical corregido"
                })

            if alertas:
                alertas_por_documento[resultado['pagina']] = alertas

        documentos_finales = []
        for idx, doc_seg in enumerate(documentos_segmentados):
            alertas_segmento = []
            for pagina in doc_seg['paginas']:
                if pagina in alertas_por_documento:
                    alertas_segmento.extend(alertas_por_documento[pagina])

            nombre_salida = f"{nombre_archivo.replace('.pdf', '')}_{doc_seg['tipo']}_{idx+1}.pdf"

            documentos_finales.append({
                "archivo_origen": nombre_archivo,
                "nombre_salida": nombre_salida,
                "tipo": doc_seg['tipo'],
                "paginas": doc_seg['paginas'],
                "pdf_bytes": doc_seg['pdf_bytes'],
                "alertas": alertas_segmento if alertas_segmento else None,
                "datos_extraidos": None
            })

        return {
            "documentos_finales": documentos_finales,
            "clasificaciones": clasificaciones,
            "error": None
        }

    except asyncio.TimeoutError:
        logger.error(f"Timeout en procesamiento de calidad para {nombre_archivo}")
        return {
            "documentos_finales": [],
            "clasificaciones": [],
            "error": f"Timeout en procesamiento de calidad ({timeout_calidad}s)"
        }
    except Exception as e:
        logger.error(f"Error en clasificar_pdf_completo para {nombre_archivo}: {str(e)}")
        return {
            "documentos_finales": [],
            "clasificaciones": [],
            "error": str(e)
        }


async def procesar_pdf_completo(pdf_bytes: bytes, nombre_archivo: str) -> Dict[str, Any]:
    """Flujo completo de procesamiento de PDF."""
    timeout_calidad = 0
    try:
        pdf_doc_temp = fitz.open(stream=pdf_bytes, filetype="pdf")
        num_paginas = len(pdf_doc_temp)
        pdf_doc_temp.close()

        timeout_calidad = calcular_timeout_calidad(num_paginas)

        logger.info(f"Iniciando procesamiento de calidad para {nombre_archivo} ({num_paginas} páginas, timeout={timeout_calidad}s)")
        start_time = time.time()

        loop = asyncio.get_event_loop()
        pdf_bytes, resultados_paso_1 = await asyncio.wait_for(
            loop.run_in_executor(
                executor,
                _procesar_calidad_sync,
                pdf_bytes
            ),
            timeout=timeout_calidad
        )

        elapsed_time = time.time() - start_time
        logger.info(f"Procesamiento de calidad completado para {nombre_archivo} en {elapsed_time:.2f}s")

        clasificaciones = await clasificar_documento_completo(pdf_bytes)

        logger.info(f"Iniciando segmentación para {nombre_archivo}")
        start_time = time.time()

        documentos_segmentados = await loop.run_in_executor(
            executor,
            segmentar_pdf,
            pdf_bytes,
            clasificaciones
        )

        elapsed_time = time.time() - start_time
        logger.info(f"Segmentación completada para {nombre_archivo} en {elapsed_time:.2f}s - {len(documentos_segmentados)} documentos")

        alertas_por_documento = {}
        for resultado in resultados_paso_1:
            alertas = []

            if resultado['escaneada']:
                if 'INCLINADA' in resultado['orientacion']:
                    alertas.append({
                        "pagina": resultado['pagina'],
                        "tipo": "inclinado",
                        "descripcion": f"Página escaneada {resultado['orientacion']}"
                    })
                elif resultado['orientacion'] not in ['NORMAL', 'SIN TEXTO', 'SIN IMAGEN']:
                    alertas.append({
                        "pagina": resultado['pagina'],
                        "tipo": "escaneado",
                        "descripcion": f"Página escaneada: {resultado['orientacion']}"
                    })

            if resultado['rotacion_formal'] != 0:
                alertas.append({
                    "pagina": resultado['pagina'],
                    "tipo": "rotado",
                    "descripcion": f"Rotación de {resultado['rotacion_formal']}° corregida"
                })

            if not resultado['escaneada'] and resultado['orientacion'] == 'ROTADA':
                alertas.append({
                    "pagina": resultado['pagina'],
                    "tipo": "rotado",
                    "descripcion": "Texto vertical corregido"
                })

            if alertas:
                alertas_por_documento[resultado['pagina']] = alertas

        documentos_finales = []
        for idx, doc_seg in enumerate(documentos_segmentados):
            alertas_segmento = []
            for pagina in doc_seg['paginas']:
                if pagina in alertas_por_documento:
                    alertas_segmento.extend(alertas_por_documento[pagina])

            nombre_salida = f"{nombre_archivo.replace('.pdf', '')}_{doc_seg['tipo']}_{idx+1}.pdf"

            tipo_documento = doc_seg['tipo']
            model_id = DOCUMENT_TYPE_TO_MODEL.get(tipo_documento)

            datos_extraidos = None

            if model_id:
                try:
                    modelo_entrenado = await verificar_modelo_entrenado(model_id)

                    if modelo_entrenado:
                        datos_extraidos = await extraer_datos_con_modelo(
                            doc_seg['pdf_bytes'],
                            model_id
                        )
                except Exception as e:
                    logger.error(f"Error extrayendo datos: {str(e)}")
                    datos_extraidos = None

            documentos_finales.append({
                "archivo_origen": nombre_archivo,
                "nombre_salida": nombre_salida,
                "tipo": doc_seg['tipo'],
                "paginas": doc_seg['paginas'],
                "pdf_bytes": doc_seg['pdf_bytes'],
                "alertas": alertas_segmento if alertas_segmento else None,
                "datos_extraidos": datos_extraidos
            })

        return {
            "documentos_finales": documentos_finales,
            "clasificaciones": clasificaciones,
            "error": None
        }

    except asyncio.TimeoutError:
        logger.error(f"Timeout en procesamiento de calidad para {nombre_archivo}")
        return {
            "documentos_finales": [],
            "clasificaciones": [],
            "error": f"Timeout en procesamiento de calidad ({timeout_calidad}s)"
        }
    except Exception as e:
        logger.error(f"Error en procesar_pdf_completo para {nombre_archivo}: {str(e)}")
        return {
            "documentos_finales": [],
            "clasificaciones": [],
            "error": str(e)
        }


def serializar_documentos_para_cache(documentos: List[Dict]) -> List[Dict]:
    """Prepara documentos para almacenamiento en caché (sin pdf_bytes)."""
    return [
        {
            "archivo_origen": doc["archivo_origen"],
            "nombre_salida": doc["nombre_salida"],
            "tipo": doc["tipo"],
            "paginas": doc["paginas"],
            "alertas": doc.get("alertas"),
            "datos_extraidos": doc.get("datos_extraidos")
        }
        for doc in documentos
    ]


# SGD Routes
@get("/consultar/{codigo_despacho:str}")
async def consultar_despacho(request: Request, codigo_despacho: str) -> dict:
    """Consulta la información del despacho y lista los documentos disponibles."""
    verify_api_token(request)
    
    if not settings.BEARER_TOKEN:
        raise HTTPException(status_code=HTTP_500_INTERNAL_SERVER_ERROR, detail="BEARER_TOKEN no configurado")
    
    datos_despacho_detalle = await consultar_despacho_detalle(codigo_despacho, settings.BEARER_TOKEN)
    
    if datos_despacho_detalle and "data" in datos_despacho_detalle:
        despacho_data = datos_despacho_detalle["data"]
        
        if isinstance(despacho_data, dict):
            codigo_visible = str(despacho_data.get("codigo", "N/A"))
            cliente = despacho_data.get("cliente", {})
            cliente_nombre = cliente.get("nombre", "N/A") if isinstance(cliente, dict) else "N/A"
            
            documentos_list = despacho_data.get("documentos", [])
            if not isinstance(documentos_list, list):
                documentos_list = []
            
            documentos_simplificados = []
            for doc in documentos_list:
                if isinstance(doc, dict):
                    tipo = doc.get("tipo", {})
                    documentos_simplificados.append({
                        "nombre": tipo.get("nombre", "Sin nombre") if isinstance(tipo, dict) else "Sin nombre",
                        "estado": doc.get("estado", "N/A"),
                        "fecha_recepcion": doc.get("fecha_recepcion", "N/A")
                    })
            
            usuarios_list = despacho_data.get("usuarios", [])
            if not isinstance(usuarios_list, list):
                usuarios_list = []
            
            pedidores = []
            jefes_operaciones = []
            
            for user in usuarios_list:
                if isinstance(user, dict):
                    role_name = user.get("role_name", "")
                    nombre = user.get("name", "")
                    
                    if role_name in ("pedidor", "pedidor_exportaciones"):
                        pedidores.append(nombre)
                    elif role_name == "jefe_operaciones":
                        jefes_operaciones.append(nombre)
            
            usuarios_datos = {
                "pedidor": pedidores if pedidores else None,
                "jefe_operaciones": jefes_operaciones if jefes_operaciones else None
            }
            
            return {
                "codigo_despacho": codigo_visible,
                "id_interno": str(despacho_data.get("id", "N/A")),
                "cliente": cliente_nombre,
                "estado": despacho_data.get("estado_despacho", "N/A"),
                "tipo": despacho_data.get("tipo_despacho", "N/A"),
                "total_documentos": len(documentos_simplificados),
                "documentos": documentos_simplificados,
                "usuarios": usuarios_datos
            }
    
    documentos_base64_list = await consultar_documentacion(codigo_despacho, settings.BEARER_TOKEN)
    
    if documentos_base64_list and isinstance(documentos_base64_list, list):
        documentos_info = []
        for doc in documentos_base64_list:
            if isinstance(doc, dict):
                documentos_info.append({
                    "nombre": doc.get("nombre_documento", "Sin nombre"),
                    "documento_id": doc.get("documento_id", "N/A")
                })
        
        return {
            "codigo": codigo_despacho,
            "mensaje": "Información limitada: solo se encontraron documentos",
            "total_documentos": len(documentos_info),
            "documentos": documentos_info
        }
    
    raise HTTPException(status_code=HTTP_404_NOT_FOUND, detail="Despacho no encontrado")


@post("/clasificar/{codigo_despacho:str}")
async def clasificar_despacho(
    request: Request, 
    codigo_despacho: str,
    force: bool = False
) -> ProcesamientoResponse:
    """
    Clasifica los documentos del despacho.
    
    Args:
        codigo_despacho: Código del despacho a clasificar
        force: Si es True, fuerza el reprocesamiento ignorando el caché
    """
    verify_api_token(request)
    
    if not settings.BEARER_TOKEN:
        raise HTTPException(status_code=HTTP_500_INTERNAL_SERVER_ERROR, detail="BEARER_TOKEN no configurado")

    datos_despacho_detalle = await consultar_despacho_detalle(codigo_despacho, settings.BEARER_TOKEN)

    codigo_visible = None
    cliente_nombre = "N/A"
    estado_despacho = "N/A"
    tipo_despacho = "N/A"

    if datos_despacho_detalle and "data" in datos_despacho_detalle:
        despacho_data = datos_despacho_detalle["data"]
        if isinstance(despacho_data, dict):
            codigo_visible = str(despacho_data.get("codigo", ""))
            cliente = despacho_data.get("cliente", {})
            cliente_nombre = cliente.get("nombre", "N/A") if isinstance(cliente, dict) else "N/A"
            estado_despacho = despacho_data.get("estado_despacho", "N/A")
            tipo_despacho = despacho_data.get("tipo_despacho", "N/A")

    documentos_base64_list = None

    if codigo_visible:
        documentos_base64_list = await consultar_documentacion(codigo_visible, settings.BEARER_TOKEN)

    if not documentos_base64_list:
        documentos_base64_list = await consultar_documentacion(codigo_despacho, settings.BEARER_TOKEN)

    if not documentos_base64_list or not isinstance(documentos_base64_list, list):
        raise HTTPException(status_code=HTTP_404_NOT_FOUND, detail="No se pudieron obtener los documentos")

    # Calcular hash de documentos actuales
    documentos_hash = calcular_hash_documentos(documentos_base64_list)

    # Verificar caché si está habilitado y no se fuerza reprocesamiento
    cache_info = None
    if settings.CACHE_ENABLED and not force:
        try:
            estado_cache = await cache_repo.verificar_cambios_despacho(
                codigo_despacho, "clasificar", documentos_hash
            )
            
            if estado_cache["existe_cache"] and not estado_cache["hay_cambios"]:
                # Retornar desde caché
                cached = await cache_repo.obtener_despacho(
                    codigo_despacho, "clasificar", documentos_hash
                )
                if cached:
                    logger.info(f"Retornando clasificación desde caché para {codigo_despacho}")
                    docs_response = [
                        DocumentoFinal(**doc) for doc in cached["resultado"]["documentos"]
                    ]
                    return ProcesamientoResponse(
                        codigo_despacho=codigo_despacho,
                        cliente=cached["cliente"],
                        estado=cached["estado"],
                        tipo=cached["tipo"],
                        total_documentos_segmentados=cached["total_documentos_segmentados"],
                        documentos=docs_response,
                        cache_info=CacheInfo(
                            desde_cache=True,
                            hash_documentos=documentos_hash,
                            fecha_cache=str(cached["updated_at"]),
                            hay_cambios=False
                        )
                    )
            
            # Hay cambios o no existe caché
            cache_info = CacheInfo(
                desde_cache=False,
                hash_documentos=documentos_hash,
                hay_cambios=estado_cache.get("hay_cambios", True)
            )
        except Exception as e:
            logger.warning(f"Error consultando caché: {e}")
            cache_info = CacheInfo(desde_cache=False, hash_documentos=documentos_hash)

    # Procesar documentos
    todos_documentos_finales = []

    for doc in documentos_base64_list:
        if isinstance(doc, dict):
            base64_data = doc.get("documento", "")

            if ',' in base64_data:
                _, base64_content = base64_data.split(',', 1)
            else:
                base64_content = base64_data

            try:
                file_bytes = base64.b64decode(base64_content)
                nombre_documento = doc.get("nombre_documento", "documento.pdf")

                validar_tamano_archivo(file_bytes)

                if es_archivo_excel(nombre_documento):
                    if not validar_excel(file_bytes, nombre_documento):
                        continue
                    try:
                        pdf_bytes = await convertir_excel_a_pdf(file_bytes, nombre_documento)
                        nombre_procesamiento = nombre_documento.rsplit('.', 1)[0] + '.pdf'
                    except Exception:
                        continue
                elif es_archivo_imagen(nombre_documento):
                    if not validar_imagen(file_bytes, nombre_documento):
                        continue
                    try:
                        pdf_bytes = await convertir_imagen_a_pdf(file_bytes, nombre_documento)
                        nombre_procesamiento = nombre_documento.rsplit('.', 1)[0] + '.pdf'
                    except Exception:
                        continue
                else:
                    if not validar_pdf(file_bytes):
                        continue
                    pdf_bytes = file_bytes
                    nombre_procesamiento = nombre_documento

                resultado = await clasificar_pdf_completo(pdf_bytes, nombre_procesamiento)

                if resultado["error"]:
                    continue

                todos_documentos_finales.extend(resultado["documentos_finales"])

            except Exception:
                continue

    documentos_finales_sgd[codigo_despacho] = todos_documentos_finales

    docs_response = []
    for doc_final in todos_documentos_finales:
        datos_limpios = None
        if doc_final.get("datos_extraidos"):
             datos_limpios = eliminar_campos_vacios(doc_final.get("datos_extraidos"))

        docs_response.append(DocumentoFinal(
            archivo_origen=doc_final["archivo_origen"],
            nombre_salida=doc_final["nombre_salida"],
            tipo=doc_final["tipo"],
            paginas=doc_final["paginas"],
            alertas=[Alerta(**alerta) for alerta in doc_final["alertas"]] if doc_final["alertas"] else None,
            datos_extraidos=datos_limpios
        ))

    # Guardar en caché
    if settings.CACHE_ENABLED:
        try:
            await cache_repo.guardar_despacho(
                codigo_despacho=codigo_despacho,
                tipo_operacion="clasificar",
                documentos_hash=documentos_hash,
                cliente=cliente_nombre,
                estado=estado_despacho,
                tipo=tipo_despacho,
                total_documentos_segmentados=len(docs_response),
                resultado={"documentos": [doc.model_dump() for doc in docs_response]}
            )
        except Exception as e:
            logger.warning(f"Error guardando en caché: {e}")

    return ProcesamientoResponse(
        codigo_despacho=codigo_despacho,
        cliente=cliente_nombre,
        estado=estado_despacho,
        tipo=tipo_despacho,
        total_documentos_segmentados=len(docs_response),
        documentos=docs_response,
        cache_info=cache_info
    )


@post("/procesar/{codigo_despacho:str}")
async def procesar_despacho(
    request: Request, 
    codigo_despacho: str,
    force: bool = False
) -> ProcesamientoResponse:
    """
    Procesa el despacho completo (clasificación + extracción de datos).
    
    Args:
        codigo_despacho: Código del despacho a procesar
        force: Si es True, fuerza el reprocesamiento ignorando el caché
    """
    verify_api_token(request)
    
    if not settings.BEARER_TOKEN:
        raise HTTPException(status_code=HTTP_500_INTERNAL_SERVER_ERROR, detail="BEARER_TOKEN no configurado")
    
    datos_despacho_detalle = await consultar_despacho_detalle(codigo_despacho, settings.BEARER_TOKEN)
    
    codigo_visible = None
    cliente_nombre = "N/A"
    estado_despacho = "N/A"
    tipo_despacho = "N/A"
    
    if datos_despacho_detalle and "data" in datos_despacho_detalle:
        despacho_data = datos_despacho_detalle["data"]
        if isinstance(despacho_data, dict):
            codigo_visible = str(despacho_data.get("codigo", ""))
            cliente = despacho_data.get("cliente", {})
            cliente_nombre = cliente.get("nombre", "N/A") if isinstance(cliente, dict) else "N/A"
            estado_despacho = despacho_data.get("estado_despacho", "N/A")
            tipo_despacho = despacho_data.get("tipo_despacho", "N/A")
    
    documentos_base64_list = None
    
    if codigo_visible:
        documentos_base64_list = await consultar_documentacion(codigo_visible, settings.BEARER_TOKEN)
    
    if not documentos_base64_list:
        documentos_base64_list = await consultar_documentacion(codigo_despacho, settings.BEARER_TOKEN)
    
    if not documentos_base64_list or not isinstance(documentos_base64_list, list):
        raise HTTPException(status_code=HTTP_404_NOT_FOUND, detail="No se pudieron obtener los documentos")
    
    # Calcular hash de documentos actuales
    documentos_hash = calcular_hash_documentos(documentos_base64_list)

    # Verificar caché si está habilitado y no se fuerza reprocesamiento
    cache_info = None
    if settings.CACHE_ENABLED and not force:
        try:
            estado_cache = await cache_repo.verificar_cambios_despacho(
                codigo_despacho, "procesar", documentos_hash
            )
            
            if estado_cache["existe_cache"] and not estado_cache["hay_cambios"]:
                cached = await cache_repo.obtener_despacho(
                    codigo_despacho, "procesar", documentos_hash
                )
                if cached:
                    logger.info(f"Retornando procesamiento desde caché para {codigo_despacho}")
                    docs_response = [
                        DocumentoFinal(**doc) for doc in cached["resultado"]["documentos"]
                    ]
                    return ProcesamientoResponse(
                        codigo_despacho=codigo_despacho,
                        cliente=cached["cliente"],
                        estado=cached["estado"],
                        tipo=cached["tipo"],
                        total_documentos_segmentados=cached["total_documentos_segmentados"],
                        documentos=docs_response,
                        cache_info=CacheInfo(
                            desde_cache=True,
                            hash_documentos=documentos_hash,
                            fecha_cache=str(cached["updated_at"]),
                            hay_cambios=False
                        )
                    )
            
            cache_info = CacheInfo(
                desde_cache=False,
                hash_documentos=documentos_hash,
                hay_cambios=estado_cache.get("hay_cambios", True)
            )
        except Exception as e:
            logger.warning(f"Error consultando caché: {e}")
            cache_info = CacheInfo(desde_cache=False, hash_documentos=documentos_hash)

    todos_documentos_finales = []
    
    for doc in documentos_base64_list:
        if isinstance(doc, dict):
            base64_data = doc.get("documento", "")
            
            if ',' in base64_data:
                _, base64_content = base64_data.split(',', 1)
            else:
                base64_content = base64_data
            
            try:
                file_bytes = base64.b64decode(base64_content)
                nombre_documento = doc.get("nombre_documento", "documento.pdf")
                
                validar_tamano_archivo(file_bytes)

                if es_archivo_excel(nombre_documento):
                    if not validar_excel(file_bytes, nombre_documento):
                        continue
                    try:
                        pdf_bytes = await convertir_excel_a_pdf(file_bytes, nombre_documento)
                        nombre_procesamiento = nombre_documento.rsplit('.', 1)[0] + '.pdf'
                    except Exception:
                        continue
                elif es_archivo_imagen(nombre_documento):
                    if not validar_imagen(file_bytes, nombre_documento):
                        continue
                    try:
                        pdf_bytes = await convertir_imagen_a_pdf(file_bytes, nombre_documento)
                        nombre_procesamiento = nombre_documento.rsplit('.', 1)[0] + '.pdf'
                    except Exception:
                        continue
                else:
                    if not validar_pdf(file_bytes):
                        continue
                    pdf_bytes = file_bytes
                    nombre_procesamiento = nombre_documento

                resultado = await procesar_pdf_completo(pdf_bytes, nombre_procesamiento)
                
                if resultado["error"]:
                    continue
                
                todos_documentos_finales.extend(resultado["documentos_finales"])
                
            except Exception:
                continue
    
    documentos_finales_sgd[codigo_despacho] = todos_documentos_finales
    
    docs_response = []
    for doc_final in todos_documentos_finales:
        datos_limpios = None
        if doc_final.get("datos_extraidos"):
             datos_limpios = eliminar_campos_vacios(doc_final.get("datos_extraidos"))

        docs_response.append(DocumentoFinal(
            archivo_origen=doc_final["archivo_origen"],
            nombre_salida=doc_final["nombre_salida"],
            tipo=doc_final["tipo"],
            paginas=doc_final["paginas"],
            alertas=[Alerta(**alerta) for alerta in doc_final["alertas"]] if doc_final["alertas"] else None,
            datos_extraidos=datos_limpios
        ))

    # Guardar en caché
    if settings.CACHE_ENABLED:
        try:
            await cache_repo.guardar_despacho(
                codigo_despacho=codigo_despacho,
                tipo_operacion="procesar",
                documentos_hash=documentos_hash,
                cliente=cliente_nombre,
                estado=estado_despacho,
                tipo=tipo_despacho,
                total_documentos_segmentados=len(docs_response),
                resultado={"documentos": [doc.model_dump() for doc in docs_response]}
            )
        except Exception as e:
            logger.warning(f"Error guardando en caché: {e}")

    return ProcesamientoResponse(
        codigo_despacho=codigo_despacho,
        cliente=cliente_nombre,
        estado=estado_despacho,
        tipo=tipo_despacho,
        total_documentos_segmentados=len(docs_response),
        documentos=docs_response,
        cache_info=cache_info
    )


# Documentos Routes
@post("/clasificar")
async def clasificar_documento_individual(
    request: Request,
    data: Annotated[UploadFile, Body(media_type=RequestEncodingType.MULTI_PART)],
    force: bool = False
) -> ProcesamientoIndividualResponse:
    """
    Clasifica un documento individual.
    
    Args:
        data: Archivo a clasificar
        force: Si es True, fuerza el reprocesamiento ignorando el caché
    """
    verify_api_token(request)
    
    nombre_archivo = data.filename.lower()
    es_pdf = nombre_archivo.endswith('.pdf')
    es_excel = es_archivo_excel(data.filename)
    es_imagen = es_archivo_imagen(data.filename)

    if not es_pdf and not es_excel and not es_imagen:
        raise HTTPException(
            status_code=HTTP_400_BAD_REQUEST,
            detail="Solo se aceptan archivos PDF, Excel (.xls, .xlsx, .xlsm, .xlsb, .xltx, .xltm) o imágenes (.jpg, .jpeg, .png, .gif, .bmp, .tiff, .webp)"
        )

    try:
        file_bytes = await data.read()
        archivo_hash = calcular_hash_archivo(file_bytes)

        validar_tamano_archivo(file_bytes)

        # Verificar caché
        cache_info = None
        if settings.CACHE_ENABLED and not force:
            try:
                cached = await cache_repo.obtener_documento(archivo_hash, "clasificar")
                if cached:
                    logger.info(f"Retornando clasificación desde caché para {data.filename}")
                    docs_response = [
                        DocumentoFinal(**doc) for doc in cached["resultado"]["documentos"]
                    ]
                    return ProcesamientoIndividualResponse(
                        archivo_origen=data.filename,
                        total_documentos_segmentados=cached["total_documentos_segmentados"],
                        documentos=docs_response,
                        cache_info=CacheInfo(
                            desde_cache=True,
                            hash_documentos=archivo_hash,
                            fecha_cache=str(cached["updated_at"])
                        )
                    )
                cache_info = CacheInfo(desde_cache=False, hash_documentos=archivo_hash)
            except Exception as e:
                logger.warning(f"Error consultando caché: {e}")
                cache_info = CacheInfo(desde_cache=False, hash_documentos=archivo_hash)

        if es_excel:
            if not validar_excel(file_bytes, data.filename):
                raise HTTPException(
                    status_code=HTTP_400_BAD_REQUEST,
                    detail="El archivo no es un Excel válido"
                )
            pdf_bytes = await convertir_excel_a_pdf(file_bytes, data.filename)
            nombre_procesamiento = data.filename.rsplit('.', 1)[0] + '.pdf'
        elif es_imagen:
            if not validar_imagen(file_bytes, data.filename):
                raise HTTPException(
                    status_code=HTTP_400_BAD_REQUEST,
                    detail="El archivo no es una imagen válida"
                )
            pdf_bytes = await convertir_imagen_a_pdf(file_bytes, data.filename)
            nombre_procesamiento = data.filename.rsplit('.', 1)[0] + '.pdf'
        else:
            if not validar_pdf(file_bytes):
                raise HTTPException(
                    status_code=HTTP_400_BAD_REQUEST,
                    detail="El archivo no es un PDF válido"
                )
            pdf_bytes = file_bytes
            nombre_procesamiento = data.filename

        resultado = await clasificar_pdf_completo(pdf_bytes, nombre_procesamiento)

        if resultado["error"]:
            raise HTTPException(status_code=HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Error al clasificar: {resultado['error']}")

        documentos_finales_individuales[data.filename] = resultado["documentos_finales"]

        docs_response = []
        for doc_final in resultado["documentos_finales"]:
            datos_limpios = None
            if doc_final.get("datos_extraidos"):
                datos_limpios = eliminar_campos_vacios(doc_final.get("datos_extraidos"))

            docs_response.append(DocumentoFinal(
                archivo_origen=doc_final["archivo_origen"],
                nombre_salida=doc_final["nombre_salida"],
                tipo=doc_final["tipo"],
                paginas=doc_final["paginas"],
                alertas=[Alerta(**alerta) for alerta in doc_final["alertas"]] if doc_final["alertas"] else None,
                datos_extraidos=datos_limpios
            ))

        # Guardar en caché
        if settings.CACHE_ENABLED:
            try:
                await cache_repo.guardar_documento(
                    archivo_hash=archivo_hash,
                    nombre_archivo=data.filename,
                    tipo_operacion="clasificar",
                    total_documentos_segmentados=len(docs_response),
                    resultado={"documentos": [doc.model_dump() for doc in docs_response]}
                )
            except Exception as e:
                logger.warning(f"Error guardando en caché: {e}")

        return ProcesamientoIndividualResponse(
            archivo_origen=data.filename,
            total_documentos_segmentados=len(docs_response),
            documentos=docs_response,
            cache_info=cache_info
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Error al clasificar documento: {str(e)}")


@post("/procesar")
async def procesar_documento_individual(
    request: Request,
    data: Annotated[UploadFile, Body(media_type=RequestEncodingType.MULTI_PART)],
    force: bool = False
) -> ProcesamientoIndividualResponse:
    """
    Procesa un documento individual completo.
    
    Args:
        data: Archivo a procesar
        force: Si es True, fuerza el reprocesamiento ignorando el caché
    """
    verify_api_token(request)
    
    nombre_archivo = data.filename.lower()
    es_pdf = nombre_archivo.endswith('.pdf')
    es_excel = es_archivo_excel(data.filename)
    es_imagen = es_archivo_imagen(data.filename)

    if not es_pdf and not es_excel and not es_imagen:
        raise HTTPException(
            status_code=HTTP_400_BAD_REQUEST,
            detail="Solo se aceptan archivos PDF, Excel (.xls, .xlsx, .xlsm, .xlsb, .xltx, .xltm) o imágenes (.jpg, .jpeg, .png, .gif, .bmp, .tiff, .webp)"
        )

    try:
        file_bytes = await data.read()
        archivo_hash = calcular_hash_archivo(file_bytes)

        validar_tamano_archivo(file_bytes)

        # Verificar caché
        cache_info = None
        if settings.CACHE_ENABLED and not force:
            try:
                cached = await cache_repo.obtener_documento(archivo_hash, "procesar")
                if cached:
                    logger.info(f"Retornando procesamiento desde caché para {data.filename}")
                    docs_response = [
                        DocumentoFinal(**doc) for doc in cached["resultado"]["documentos"]
                    ]
                    return ProcesamientoIndividualResponse(
                        archivo_origen=data.filename,
                        total_documentos_segmentados=cached["total_documentos_segmentados"],
                        documentos=docs_response,
                        cache_info=CacheInfo(
                            desde_cache=True,
                            hash_documentos=archivo_hash,
                            fecha_cache=str(cached["updated_at"])
                        )
                    )
                cache_info = CacheInfo(desde_cache=False, hash_documentos=archivo_hash)
            except Exception as e:
                logger.warning(f"Error consultando caché: {e}")
                cache_info = CacheInfo(desde_cache=False, hash_documentos=archivo_hash)

        if es_excel:
            if not validar_excel(file_bytes, data.filename):
                raise HTTPException(
                    status_code=HTTP_400_BAD_REQUEST,
                    detail="El archivo no es un Excel válido"
                )
            pdf_bytes = await convertir_excel_a_pdf(file_bytes, data.filename)
            nombre_procesamiento = data.filename.rsplit('.', 1)[0] + '.pdf'
        elif es_imagen:
            if not validar_imagen(file_bytes, data.filename):
                raise HTTPException(
                    status_code=HTTP_400_BAD_REQUEST,
                    detail="El archivo no es una imagen válida"
                )
            pdf_bytes = await convertir_imagen_a_pdf(file_bytes, data.filename)
            nombre_procesamiento = data.filename.rsplit('.', 1)[0] + '.pdf'
        else:
            if not validar_pdf(file_bytes):
                raise HTTPException(
                    status_code=HTTP_400_BAD_REQUEST,
                    detail="El archivo no es un PDF válido"
                )
            pdf_bytes = file_bytes
            nombre_procesamiento = data.filename

        resultado = await procesar_pdf_completo(pdf_bytes, nombre_procesamiento)
        
        if resultado["error"]:
            raise HTTPException(status_code=HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Error al procesar: {resultado['error']}")
        
        documentos_finales_individuales[data.filename] = resultado["documentos_finales"]
        
        docs_response = []
        for doc_final in resultado["documentos_finales"]:
            datos_limpios = None
            if doc_final.get("datos_extraidos"):
                datos_limpios = eliminar_campos_vacios(doc_final.get("datos_extraidos"))

            docs_response.append(DocumentoFinal(
                archivo_origen=doc_final["archivo_origen"],
                nombre_salida=doc_final["nombre_salida"],
                tipo=doc_final["tipo"],
                paginas=doc_final["paginas"],
                alertas=[Alerta(**alerta) for alerta in doc_final["alertas"]] if doc_final["alertas"] else None,
                datos_extraidos=datos_limpios
            ))

        # Guardar en caché
        if settings.CACHE_ENABLED:
            try:
                await cache_repo.guardar_documento(
                    archivo_hash=archivo_hash,
                    nombre_archivo=data.filename,
                    tipo_operacion="procesar",
                    total_documentos_segmentados=len(docs_response),
                    resultado={"documentos": [doc.model_dump() for doc in docs_response]}
                )
            except Exception as e:
                logger.warning(f"Error guardando en caché: {e}")

        return ProcesamientoIndividualResponse(
            archivo_origen=data.filename,
            total_documentos_segmentados=len(docs_response),
            documentos=docs_response,
            cache_info=cache_info
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Error al procesar documento: {str(e)}")


# Admin Routes - Tokens
@get("/")
async def listar_tokens(request: Request) -> List[TokenInfo]:
    """Lista todos los tokens de la API con su metadata."""
    verify_admin_token(request)
    
    tokens = token_manager.list_tokens()

    env_tokens = get_valid_api_tokens()
    for idx, token_value in enumerate(env_tokens, start=1):
        masked_token = f"{token_value[:8]}...{token_value[-4:]}" if len(token_value) > 12 else "***"

        tokens.append({
            "id": f"env-token-{idx}",
            "name": f"Token de .env #{idx}",
            "masked_token": masked_token,
            "created_at": "N/A",
            "created_by": "env-config",
            "last_used": None,
            "is_active": True
        })

    return [TokenInfo(**t) for t in tokens]


@post("/generate")
async def generar_token(request: Request, data: TokenCreateRequest) -> TokenCreateResponse:
    """Genera un nuevo token de autenticación API."""
    verify_admin_token(request)
    
    result = token_manager.generate_token(
        name=data.name,
        created_by="admin"
    )
    return TokenCreateResponse(**result)


@delete("/{token_id:str}", status_code=HTTP_200_OK)
async def eliminar_token(request: Request, token_id: str) -> TokenDeleteResponse:
    """Elimina un token por su ID."""
    verify_admin_token(request)
    
    success = token_manager.delete_token(token_id)

    if success:
        return TokenDeleteResponse(
            success=True,
            message=f"Token {token_id} eliminado exitosamente"
        )
    else:
        raise HTTPException(
            status_code=HTTP_404_NOT_FOUND,
            detail=f"Token {token_id} no encontrado"
        )


# Cache Management Routes
@delete("/cache/despacho/{codigo_despacho:str}", status_code=HTTP_200_OK)
async def eliminar_cache_despacho(
    request: Request, 
    codigo_despacho: str,
    tipo_operacion: Optional[str] = None
) -> dict:
    """Elimina el caché de un despacho específico."""
    verify_admin_token(request)
    
    try:
        count = await cache_repo.eliminar_cache_despacho(codigo_despacho, tipo_operacion)
        return {
            "success": True,
            "message": f"Eliminados {count} registros de caché",
            "codigo_despacho": codigo_despacho
        }
    except Exception as e:
        raise HTTPException(
            status_code=HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error eliminando caché: {str(e)}"
        )


@get("/")
async def root() -> dict:
    return {"status": "online", "service": "API Docs", "azure_di": "cloud", "cache": "postgresql"}


# Lifecycle events
async def on_startup():
    """Inicializa conexiones al iniciar la aplicación."""
    try:
        await db_manager.initialize()
        logger.info("Base de datos inicializada correctamente")
    except Exception as e:
        logger.error(f"Error inicializando base de datos: {e}")


async def on_shutdown():
    """Cierra conexiones al detener la aplicación."""
    await db_manager.close()
    logger.info("Conexiones cerradas")


# Crear Routers con prefijos y tags
sgd_router = Router(
    path="/sgd",
    route_handlers=[consultar_despacho, clasificar_despacho, procesar_despacho],
    tags=["SGD"]
)

documentos_router = Router(
    path="/documentos",
    route_handlers=[clasificar_documento_individual, procesar_documento_individual],
    tags=["Documentos"]
)

admin_router = Router(
    path="/admin/tokens",
    route_handlers=[listar_tokens, generar_token, eliminar_token],
    tags=["Admin - Gestión de Tokens"]
)

cache_router = Router(
    path="/admin",
    route_handlers=[eliminar_cache_despacho],
    tags=["Admin - Gestión de Caché"]
)

app = Litestar(
    route_handlers=[
        root,
        sgd_router,
        documentos_router,
        admin_router,
        cache_router,
    ],
    on_startup=[on_startup],
    on_shutdown=[on_shutdown],
    openapi_config=OpenAPIConfig(
        title="API Docs",
        version="2.1.0",
        components=Components(
            security_schemes={
                "BearerAuth": SecurityScheme(
                    type="http",
                    scheme="bearer",
                    bearer_format="JWT",
                    description="Token de autenticación API"
                )
            }
        ),
        security=[{"BearerAuth": []}]
    ),
    request_max_body_size=500 * 1024 * 1024,  # 500MB
)