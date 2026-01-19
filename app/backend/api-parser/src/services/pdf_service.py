import os
import io
import time
import asyncio
import pandas as pd
from PIL import Image
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from litestar.exceptions import HTTPException
from litestar.status_codes import HTTP_400_BAD_REQUEST, HTTP_408_REQUEST_TIMEOUT
import logging
from concurrent.futures import ThreadPoolExecutor

from config.settings import get_settings, calcular_timeout_excel

logger = logging.getLogger(__name__)
settings = get_settings()

# Setup fonts
try:
    pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
    FUENTE_PRINCIPAL = 'STSong-Light'
except Exception:
    FUENTE_PRINCIPAL = 'Helvetica'

# Executor (Normally passed or shared, but recreated here for simplicity if not injected)
# Ideally we should use a shared executor
executor = ThreadPoolExecutor(max_workers=min(settings.EXECUTOR_MAX_WORKERS, (os.cpu_count() or 1) * 4))


def formatear_celda(valor):
    """Formatea el valor de una celda para su representación en el PDF."""
    if pd.isna(valor):
        return ""

    if isinstance(valor, (float, int)):
        try:
            val_float = float(valor)
            if val_float.is_integer():
                return str(int(val_float))
            return str(val_float)
        except Exception:
            return str(valor)

    if isinstance(valor, pd.Timestamp):
        return valor.strftime('%Y-%m-%d')

    return str(valor)


def limpiar_dataframe(df):
    """Limpia el DataFrame eliminando columnas y filas vacías, detectando gaps."""
    df = df.fillna('')
    df = df.astype(str)
    df = df.replace('nan', '')

    densidades = []
    for col in df.columns:
        celdas_con_dato = (df[col].str.strip().str.len() > 0).sum()
        densidades.append(celdas_con_dato)

    umbral = 2

    hay_gap = False
    gap_actual = 0
    for d in densidades:
        if d < umbral:
            gap_actual += 1
            if gap_actual >= 3:
                hay_gap = True
                break
        else:
            gap_actual = 0

    if hay_gap:
        cols_con_datos_idx = [i for i, d in enumerate(densidades) if d >= umbral]

        if not cols_con_datos_idx:
            return pd.DataFrame()

        primer_col = cols_con_datos_idx[0]
        fin_bloque = cols_con_datos_idx[-1]

        for i in range(primer_col, len(densidades)):
            gap = 0
            for j in range(i, min(i + 5, len(densidades))):
                if densidades[j] < umbral:
                    gap += 1
                else:
                    break
            if gap >= 3:
                fin_bloque = i - 1
                break

        cols_seleccionadas = [i for i in range(primer_col, fin_bloque + 1) if densidades[i] >= umbral]
        if not cols_seleccionadas:
            return pd.DataFrame()
        df = df.iloc[:, cols_seleccionadas]
    else:
        cols_con_datos = []
        for col in df.columns:
            if df[col].str.strip().str.len().sum() > 0:
                cols_con_datos.append(col)
        df = df[cols_con_datos]

    df = df[df.apply(lambda row: row.str.strip().str.len().sum() > 0, axis=1)]

    return df.reset_index(drop=True)


def calcular_anchos_columnas_mejorado(df, disponible_width):
    """Calcula anchos de columnas óptimos basados en el contenido."""
    if df.empty:
        return []

    num_cols = len(df.columns)
    max_lens = []

    for col_idx in range(num_cols):
        col_data = df.iloc[:, col_idx].astype(str)
        longitudes = col_data.str.len()
        p90 = longitudes.quantile(0.9) if len(longitudes) > 0 else 0
        max_lens.append(max(p90, 5))

    total_chars = sum(max_lens)
    if total_chars == 0:
        total_chars = 1

    col_widths = [(l / total_chars) * disponible_width for l in max_lens]

    if num_cols <= 6:
        min_width = 0.8 * inch
        max_width = 5.0 * inch
    elif num_cols <= 10:
        min_width = 0.5 * inch
        max_width = 4.0 * inch
    else:
        min_width = 0.4 * inch
        max_width = 3.0 * inch

    col_widths = [max(min(w, max_width), min_width) for w in col_widths]

    current_total = sum(col_widths)
    if current_total > 0:
        factor = disponible_width / current_total
        col_widths = [w * factor for w in col_widths]

    return col_widths


def _convertir_imagen_a_pdf_sync(imagen_bytes: bytes, nombre_archivo: str) -> bytes:
    """Función síncrona interna para conversión de imagen a PDF."""
    try:
        imagen = Image.open(io.BytesIO(imagen_bytes))

        if imagen.mode == 'RGBA':
            fondo = Image.new('RGB', imagen.size, (255, 255, 255))
            fondo.paste(imagen, mask=imagen.split()[3])
            imagen = fondo
        elif imagen.mode != 'RGB':
            imagen = imagen.convert('RGB')

        ancho_img, alto_img = imagen.size
        ancho_pagina, alto_pagina = A4
        margen = 30

        ancho_disponible = ancho_pagina - (2 * margen)
        alto_disponible = alto_pagina - (2 * margen)

        escala = min(ancho_disponible / ancho_img, alto_disponible / alto_img)

        ancho_final = ancho_img * escala
        alto_final = alto_img * escala

        x = (ancho_pagina - ancho_final) / 2
        y = (alto_pagina - alto_final) / 2

        buffer = io.BytesIO()
        c = canvas.Canvas(buffer, pagesize=A4)

        img_reader = ImageReader(imagen)
        c.drawImage(img_reader, x, y, width=ancho_final, height=alto_final, preserveAspectRatio=True)
        c.save()

        return buffer.getvalue()

    except Exception as e:
        raise HTTPException(
            status_code=HTTP_400_BAD_REQUEST,
            detail=f"Error al convertir imagen a PDF: {str(e)}"
        )


def _convertir_excel_a_pdf_sync(excel_bytes: bytes, nombre_archivo: str) -> bytes:
    """Función síncrona interna para conversión Excel a PDF con soporte avanzado."""
    try:
        extension = os.path.splitext(nombre_archivo.lower())[1]

        df_dict = None
        errores = []

        if extension in ['.xlsx', '.xlsm']:
            try:
                df_dict = pd.read_excel(io.BytesIO(excel_bytes), sheet_name=None, engine='openpyxl', header=None)
            except Exception as e1:
                errores.append(f"openpyxl: {str(e1)[:100]}")
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)
                    df_dict = {}
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        data = []
                        for row in ws.iter_rows(values_only=True):
                            data.append(list(row))
                        if data:
                            df_dict[sheet_name] = pd.DataFrame(data)
                except Exception as e2:
                    errores.append(f"openpyxl manual: {str(e2)[:100]}")
        elif extension == '.xls':
            try:
                df_dict = pd.read_excel(io.BytesIO(excel_bytes), sheet_name=None, engine='xlrd', header=None)
            except Exception as e:
                errores.append(f"xlrd: {str(e)[:100]}")
        else:
            try:
                df_dict = pd.read_excel(io.BytesIO(excel_bytes), sheet_name=None, engine='openpyxl', header=None)
            except Exception as e:
                errores.append(f"openpyxl: {str(e)[:100]}")

        if not df_dict:
            logger.warning(f"No se pudo leer Excel {nombre_archivo}: {'; '.join(errores)}")
            raise HTTPException(
                status_code=HTTP_400_BAD_REQUEST,
                detail=f"Error al leer Excel: {'; '.join(errores)}"
            )

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=0.3*inch,
            rightMargin=0.3*inch,
            topMargin=0.4*inch,
            bottomMargin=0.4*inch
        )

        elements = []
        styles = getSampleStyleSheet()

        estilo_titulo = ParagraphStyle(
            'TituloHoja',
            parent=styles['Heading1'],
            fontName=FUENTE_PRINCIPAL,
            fontSize=14,
            spaceAfter=12
        )

        estilo_celda = ParagraphStyle(
            'CeldaTabla',
            fontName=FUENTE_PRINCIPAL,
            fontSize=9,
            leading=11,
        )

        for sheet_name, df in df_dict.items():
            elements.append(Paragraph(f"Hoja: {sheet_name}", estilo_titulo))

            df_clean = limpiar_dataframe(df)

            if df_clean.empty:
                continue

            data = []
            for row in df_clean.values.tolist():
                fila_procesada = []
                for c in row:
                    texto = formatear_celda(c)
                    fila_procesada.append(Paragraph(texto, estilo_celda))
                data.append(fila_procesada)

            if not data:
                continue

            ancho_util = landscape(A4)[0] - 0.6*inch
            col_widths = calcular_anchos_columnas_mejorado(df_clean, ancho_util)

            table = Table(data, colWidths=col_widths, repeatRows=1)

            estilo_grid = [
                ('FONTNAME', (0, 0), (-1, -1), FUENTE_PRINCIPAL),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.93, 0.93, 0.93)),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ]

            table.setStyle(TableStyle(estilo_grid))

            elements.append(table)
            elements.append(PageBreak())

        doc.build(elements)
        return buffer.getvalue()

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=HTTP_400_BAD_REQUEST,
            detail=f"Error al convertir Excel a PDF: {str(e)}"
        )


async def convertir_excel_a_pdf(excel_bytes: bytes, nombre_archivo: str) -> bytes:
    """Convierte un archivo Excel a PDF de forma asíncrona con timeout dinámico."""
    timeout = calcular_timeout_excel(len(excel_bytes))
    file_size_mb = len(excel_bytes) / (1024 * 1024)

    logger.info(f"Iniciando conversión Excel a PDF para {nombre_archivo} ({file_size_mb:.2f}MB, timeout={timeout}s)")
    start_time = time.time()

    loop = asyncio.get_event_loop()

    try:
        result = await asyncio.wait_for(
            loop.run_in_executor(
                executor,
                _convertir_excel_a_pdf_sync,
                excel_bytes,
                nombre_archivo
            ),
            timeout=timeout
        )

        elapsed_time = time.time() - start_time
        logger.info(f"Conversión Excel completada para {nombre_archivo} en {elapsed_time:.2f}s")

        return result
    except asyncio.TimeoutError:
        logger.error(f"Conversión Excel excedió timeout para {nombre_archivo} ({timeout}s)")
        raise HTTPException(
            status_code=HTTP_408_REQUEST_TIMEOUT,
            detail=f"Conversión Excel excedió el tiempo límite de {timeout}s"
        )


async def convertir_imagen_a_pdf(imagen_bytes: bytes, nombre_archivo: str) -> bytes:
    """Convierte una imagen a PDF de forma asíncrona."""
    file_size_mb = len(imagen_bytes) / (1024 * 1024)
    timeout = 60

    logger.info(f"Iniciando conversión imagen a PDF para {nombre_archivo} ({file_size_mb:.2f}MB)")
    start_time = time.time()

    loop = asyncio.get_event_loop()

    try:
        result = await asyncio.wait_for(
            loop.run_in_executor(
                executor,
                _convertir_imagen_a_pdf_sync,
                imagen_bytes,
                nombre_archivo
            ),
            timeout=timeout
        )

        elapsed_time = time.time() - start_time
        logger.info(f"Conversión imagen completada para {nombre_archivo} en {elapsed_time:.2f}s")

        return result
    except asyncio.TimeoutError:
        logger.error(f"Conversión imagen excedió timeout para {nombre_archivo} ({timeout}s)")
        raise HTTPException(
            status_code=HTTP_408_REQUEST_TIMEOUT,
            detail=f"Conversión imagen excedió el tiempo límite de {timeout}s"
        )
